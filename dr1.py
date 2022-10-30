import openpyxl
book = openpyxl.load_workbook("D:\\sweetin aalesta\\practice\\p1.xltx")
sheet = book.active

dict={}

for i in range(1,sheet.max_row+1):#+1 is for get the last value (<=)
    #to get the only one row
    if sheet.cell(row=i, column=1).value == "ts2":
        for j in range(2,sheet.max_column+1):
            dict[sheet.cell(row=1, column=j).value] =sheet.cell(row=i, column=j).value

print(dict)