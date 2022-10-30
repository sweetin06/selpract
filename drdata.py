import openpyxl
book = openpyxl.load_workbook("D:\\sweetin aalesta\\practice\\p1.xltx")
sheet = book.active #active
#python starts from row 1, column 1
cell = sheet.cell(row=1, column=2)
print(cell.value)
sheet.cell(row=2, column=2).value="sweetins"

print(sheet.cell(row=2, column=2).value)
print(sheet.max_row)
print(sheet.max_column)


print(sheet['C1'].value)

for i in range(1,sheet.max_row+1):#+1 is for get the last value (<=)
    #to get the only one row
    if sheet.cell(row=i, column=1).value == "ts2":
        for j in range(2,sheet.max_column+1):
            print("value", sheet.cell(row=i, column=j).value)